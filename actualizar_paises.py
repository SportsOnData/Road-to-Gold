from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------
# Coloca este archivo en la carpeta principal del proyecto.
# Estructura esperada:
#
# Road to Gold/
# ├─ actualizar_paises.py
# ├─ paises.xlsx
# └─ datos/
#    └─ paises.js
#
BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "datos" / "paises.xlsx"
OUTPUT_PATH = BASE_DIR / "datos" / "paises.js"
SHEET_NAME: str | None = None  # None = primera hoja del Excel

COLUMNAS_OBLIGATORIAS = {"id", "codigo", "nombre", "continente", "activo"}
PREFIJO_NOTA = "nota"
NOTA_MINIMA = 0
NOTA_MAXIMA = 10


def convertir_booleano(valor: Any, fila: int) -> bool:
    """Convierte TRUE/FALSE, true/false, 1/0, sí/no en booleano real."""
    if isinstance(valor, bool):
        return valor

    if isinstance(valor, (int, float)) and valor in (0, 1):
        return bool(valor)

    if isinstance(valor, str):
        texto = valor.strip().lower()
        if texto in {"true", "verdadero", "sí", "si", "1"}:
            return True
        if texto in {"false", "falso", "no", "0"}:
            return False

    raise ValueError(
        f"Fila {fila}: el valor de 'activo' debe ser TRUE/FALSE "
        f"(también se aceptan true/false, sí/no o 1/0). Valor recibido: {valor!r}"
    )


def convertir_entero(valor: Any, columna: str, fila: int) -> int:
    """Convierte un valor numérico de Excel en entero y valida que no tenga decimales."""
    if valor is None or valor == "":
        raise ValueError(f"Fila {fila}: la columna '{columna}' está vacía.")

    try:
        numero = float(valor)
    except (TypeError, ValueError) as error:
        raise ValueError(
            f"Fila {fila}: '{columna}' debe ser un número. Valor recibido: {valor!r}"
        ) from error

    if not numero.is_integer():
        raise ValueError(
            f"Fila {fila}: '{columna}' debe ser un entero. Valor recibido: {valor!r}"
        )

    return int(numero)


def limpiar_texto(valor: Any, columna: str, fila: int) -> str:
    if valor is None:
        raise ValueError(f"Fila {fila}: la columna '{columna}' está vacía.")

    texto = str(valor).strip()
    if not texto:
        raise ValueError(f"Fila {fila}: la columna '{columna}' está vacía.")

    return texto


def codigo_deporte_desde_columna(columna: str) -> str:
    """
    Convierte:
      notaAtletismo -> atletismo
      notaNatacion  -> natacion
      notaTenismesa -> tenismesa
    """
    nombre = columna[len(PREFIJO_NOTA):]
    if not nombre:
        raise ValueError(f"Columna de nota inválida: {columna!r}")

    return nombre[0].lower() + nombre[1:]


def leer_paises() -> list[dict[str, Any]]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"No se encuentra el Excel:\n{EXCEL_PATH}\n\n"
            "Renombra tu archivo como 'paises.xlsx' y colócalo junto a este programa."
        )

    # data_only=True permite leer el resultado guardado de fórmulas como =A5+1.
    workbook = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    worksheet = workbook[SHEET_NAME] if SHEET_NAME else workbook.active

    filas = worksheet.iter_rows(values_only=True)

    try:
        encabezados_brutos = next(filas)
    except StopIteration as error:
        raise ValueError("El Excel está vacío.") from error

    encabezados = [
        str(valor).strip() if valor is not None else ""
        for valor in encabezados_brutos
    ]

    if len(encabezados) != len(set(encabezados)):
        repetidos = sorted(
            {nombre for nombre in encabezados if encabezados.count(nombre) > 1}
        )
        raise ValueError(f"Hay columnas repetidas en el Excel: {', '.join(repetidos)}")

    faltantes = COLUMNAS_OBLIGATORIAS - set(encabezados)
    if faltantes:
        raise ValueError(
            "Faltan columnas obligatorias: " + ", ".join(sorted(faltantes))
        )

    columnas_nota = [
        nombre for nombre in encabezados
        if nombre.startswith(PREFIJO_NOTA) and len(nombre) > len(PREFIJO_NOTA)
    ]

    if not columnas_nota:
        raise ValueError(
            "No se ha encontrado ninguna columna de notas "
            "(por ejemplo: notaAtletismo, notaNatacion...)."
        )

    indice = {nombre: posicion for posicion, nombre in enumerate(encabezados)}
    paises: list[dict[str, Any]] = []
    ids_usados: set[int] = set()
    codigos_usados: set[str] = set()

    for numero_fila, valores in enumerate(filas, start=2):
        # Ignora filas completamente vacías.
        if all(valor is None or valor == "" for valor in valores):
            continue

        # Completa filas cortas para poder acceder por índice sin error.
        valores = tuple(valores) + (None,) * (len(encabezados) - len(valores))

        identificador = convertir_entero(
            valores[indice["id"]], "id", numero_fila
        )
        codigo = limpiar_texto(
            valores[indice["codigo"]], "codigo", numero_fila
        ).upper()
        nombre = limpiar_texto(
            valores[indice["nombre"]], "nombre", numero_fila
        )
        continente = limpiar_texto(
            valores[indice["continente"]], "continente", numero_fila
        )
        activo = convertir_booleano(
            valores[indice["activo"]], numero_fila
        )

        if identificador in ids_usados:
            raise ValueError(f"Fila {numero_fila}: id repetido: {identificador}")
        if codigo in codigos_usados:
            raise ValueError(f"Fila {numero_fila}: código repetido: {codigo}")

        ids_usados.add(identificador)
        codigos_usados.add(codigo)

        notas: dict[str, int] = {}

        for columna in columnas_nota:
            nota = convertir_entero(
                valores[indice[columna]], columna, numero_fila
            )

            if not NOTA_MINIMA <= nota <= NOTA_MAXIMA:
                raise ValueError(
                    f"Fila {numero_fila}: '{columna}' debe estar entre "
                    f"{NOTA_MINIMA} y {NOTA_MAXIMA}. Valor recibido: {nota}"
                )

            codigo_deporte = codigo_deporte_desde_columna(columna)
            notas[codigo_deporte] = nota

        paises.append(
            {
                "id": identificador,
                "codigo": codigo,
                "nombre": nombre,
                "continente": continente,
                "activo": activo,
                "notas": notas,
            }
        )

    if not paises:
        raise ValueError("No se ha encontrado ningún país en el Excel.")

    return paises


def escribir_javascript(paises: list[dict[str, Any]]) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # Guarda una copia de seguridad del JS anterior.
    if OUTPUT_PATH.exists():
        backup_path = OUTPUT_PATH.with_suffix(".js.bak")
        shutil.copy2(OUTPUT_PATH, backup_path)

    contenido_json = json.dumps(
        paises,
        ensure_ascii=False,
        indent=2,
    )

    contenido_js = (
        "// Archivo generado automáticamente desde paises.xlsx.\n"
        "// No edites este archivo a mano: ejecuta actualizar_paises.py.\n"
        "// Las notas se consultan con pais.notas[deporte.codigo].\n\n"
        f"const PAISES = {contenido_json};\n"
    )

    archivo_temporal = OUTPUT_PATH.with_suffix(".js.tmp")
    archivo_temporal.write_text(contenido_js, encoding="utf-8")
    archivo_temporal.replace(OUTPUT_PATH)


def main() -> None:
    try:
        paises = leer_paises()
        escribir_javascript(paises)

        activos = sum(1 for pais in paises if pais["activo"])
        inactivos = len(paises) - activos

        print("✅ paises.js actualizado correctamente.")
        print(f"   Países totales: {len(paises)}")
        print(f"   Activos: {activos}")
        print(f"   Inactivos: {inactivos}")
        print(f"   Archivo creado: {OUTPUT_PATH}")

        backup_path = OUTPUT_PATH.with_suffix(".js.bak")
        if backup_path.exists():
            print(f"   Copia anterior: {backup_path}")

    except Exception as error:
        print("\n❌ No se pudo actualizar paises.js:")
        print(f"   {error}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
