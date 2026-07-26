from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------
# Estructura esperada:
#
# Road to Gold/
# ├─ actualizar_deportes.py
# └─ datos/
#    ├─ deportes.xlsx
#    └─ deportes.js
#
BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "datos" / "deportes.xlsx"
OUTPUT_PATH = BASE_DIR / "datos" / "deportes.js"
SHEET_NAME: str | None = None  # None = primera hoja del Excel

COLUMNAS_OBLIGATORIAS = {
    "id",
    "codigo",
    "nombre",
    "grupo",
    "tamanoSimulacion",
    "variabilidad",
    "pesoAparicion",
    "permiteEventos",
    "activo",
    "icono",
    "medallasPorPrueba",
    "categoria",
}

GRUPOS_VALIDOS = {"fijo", "fijo_alternativo", "grande", "mediano", "pequeno"}


def convertir_booleano(valor: Any, columna: str, fila: int) -> bool:
    """Convierte TRUE/FALSE, true/false, 1/0 y sí/no en booleano real."""
    if isinstance(valor, bool):
        return valor

    if isinstance(valor, (int, float)) and valor in (0, 1):
        return bool(valor)

    if isinstance(valor, str):
        texto = valor.strip().lower()

        if texto in {"true", "verdadero", "sí", "si", "1"}:
            return True

        if texto in {"false", "falso", "no", "0"}:
            return False

    raise ValueError(
        f"Fila {fila}: el valor de '{columna}' debe ser TRUE/FALSE "
        f"(también se aceptan true/false, sí/no o 1/0). "
        f"Valor recibido: {valor!r}"
    )


def convertir_entero(
    valor: Any,
    columna: str,
    fila: int,
    minimo: int | None = None,
    maximo: int | None = None,
) -> int:
    """Convierte un valor numérico en entero y valida sus límites."""
    if valor is None or valor == "":
        raise ValueError(f"Fila {fila}: la columna '{columna}' está vacía.")

    try:
        numero = float(valor)
    except (TypeError, ValueError) as error:
        raise ValueError(
            f"Fila {fila}: '{columna}' debe ser un número. "
            f"Valor recibido: {valor!r}"
        ) from error

    if not numero.is_integer():
        raise ValueError(
            f"Fila {fila}: '{columna}' debe ser un entero. "
            f"Valor recibido: {valor!r}"
        )

    entero = int(numero)

    if minimo is not None and entero < minimo:
        raise ValueError(
            f"Fila {fila}: '{columna}' no puede ser menor que {minimo}. "
            f"Valor recibido: {entero}"
        )

    if maximo is not None and entero > maximo:
        raise ValueError(
            f"Fila {fila}: '{columna}' no puede ser mayor que {maximo}. "
            f"Valor recibido: {entero}"
        )

    return entero


def limpiar_texto(valor: Any, columna: str, fila: int) -> str:
    if valor is None:
        raise ValueError(f"Fila {fila}: la columna '{columna}' está vacía.")

    texto = str(valor).strip()

    if not texto:
        raise ValueError(f"Fila {fila}: la columna '{columna}' está vacía.")

    return texto


def convertir_icono(valor: Any) -> str | None:
    """
    Una celda vacía genera null en JavaScript.
    Si contiene texto, se conserva.
    """
    if valor is None:
        return None

    texto = str(valor).strip()

    if not texto or texto.lower() in {"null", "none"}:
        return None

    return texto


def leer_deportes() -> list[dict[str, Any]]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"No se encuentra el Excel:\n{EXCEL_PATH}\n\n"
            "Renombra tu archivo como 'deportes.xlsx' y colócalo "
            "dentro de la carpeta 'datos'."
        )

    workbook = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    worksheet = workbook[SHEET_NAME] if SHEET_NAME else workbook.active

    filas = worksheet.iter_rows(values_only=True)

    try:
        encabezados_brutos = next(filas)
    except StopIteration as error:
        raise ValueError("El Excel está vacío.") from error

    encabezados = [
        str(valor).strip() if valor is not None else ""
        for valor in encabezados_brutos
    ]

    if len(encabezados) != len(set(encabezados)):
        repetidos = sorted(
            {nombre for nombre in encabezados if encabezados.count(nombre) > 1}
        )
        raise ValueError(
            "Hay columnas repetidas en el Excel: " + ", ".join(repetidos)
        )

    faltantes = COLUMNAS_OBLIGATORIAS - set(encabezados)

    if faltantes:
        raise ValueError(
            "Faltan columnas obligatorias: " + ", ".join(sorted(faltantes))
        )

    indice = {nombre: posicion for posicion, nombre in enumerate(encabezados)}

    deportes: list[dict[str, Any]] = []
    ids_usados: set[int] = set()
    codigos_usados: set[str] = set()

    for numero_fila, valores in enumerate(filas, start=2):
        if all(valor is None or valor == "" for valor in valores):
            continue

        valores = tuple(valores) + (None,) * (len(encabezados) - len(valores))

        identificador = convertir_entero(
            valores[indice["id"]],
            "id",
            numero_fila,
            minimo=1,
        )

        codigo = limpiar_texto(
            valores[indice["codigo"]],
            "codigo",
            numero_fila,
        ).lower()

        nombre = limpiar_texto(
            valores[indice["nombre"]],
            "nombre",
            numero_fila,
        )

        grupo = limpiar_texto(
            valores[indice["grupo"]],
            "grupo",
            numero_fila,
        ).lower()

        if grupo not in GRUPOS_VALIDOS:
            raise ValueError(
                f"Fila {numero_fila}: grupo no válido: {grupo!r}. "
                f"Valores admitidos: {', '.join(sorted(GRUPOS_VALIDOS))}"
            )

        tamano_simulacion = convertir_entero(
            valores[indice["tamanoSimulacion"]],
            "tamanoSimulacion",
            numero_fila,
            minimo=1,
        )

        variabilidad = convertir_entero(
            valores[indice["variabilidad"]],
            "variabilidad",
            numero_fila,
            minimo=1,
            maximo=5,
        )

        peso_aparicion = convertir_entero(
            valores[indice["pesoAparicion"]],
            "pesoAparicion",
            numero_fila,
            minimo=0,
        )

        permite_eventos = convertir_booleano(
            valores[indice["permiteEventos"]],
            "permiteEventos",
            numero_fila,
        )

        activo = convertir_booleano(
            valores[indice["activo"]],
            "activo",
            numero_fila,
        )

        icono = convertir_icono(
            valores[indice["icono"]]
        )

        medallas_por_prueba = convertir_entero(
            valores[indice["medallasPorPrueba"]],
            "medallasPorPrueba",
            numero_fila,
            minimo=1,
        )

        categoria = limpiar_texto(
            valores[indice["categoria"]],
            "categoria",
            numero_fila,
        ).lower()

        if identificador in ids_usados:
            raise ValueError(
                f"Fila {numero_fila}: id repetido: {identificador}"
            )

        if codigo in codigos_usados:
            raise ValueError(
                f"Fila {numero_fila}: código repetido: {codigo}"
            )

        ids_usados.add(identificador)
        codigos_usados.add(codigo)

        deportes.append(
            {
                "id": identificador,
                "codigo": codigo,
                "nombre": nombre,
                "grupo": grupo,
                "tamanoSimulacion": tamano_simulacion,
                "variabilidad": variabilidad,
                "pesoAparicion": peso_aparicion,
                "permiteEventos": permite_eventos,
                "activo": activo,
                "icono": icono,
                "medallasPorPrueba": medallas_por_prueba,
                "categoria": categoria,
            }
        )

    if not deportes:
        raise ValueError("No se ha encontrado ningún deporte en el Excel.")

    return deportes


def escribir_javascript(deportes: list[dict[str, Any]]) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    if OUTPUT_PATH.exists():
        backup_path = OUTPUT_PATH.with_suffix(".js.bak")
        shutil.copy2(OUTPUT_PATH, backup_path)

    contenido_json = json.dumps(
        deportes,
        ensure_ascii=False,
        indent=2,
    )

    contenido_js = (
        "// Archivo generado automáticamente desde deportes.xlsx.\n"
        "// No edites este archivo a mano: ejecuta actualizar_deportes.py.\n\n"
        f"const DEPORTES = {contenido_json};\n"
    )

    archivo_temporal = OUTPUT_PATH.with_suffix(".js.tmp")
    archivo_temporal.write_text(contenido_js, encoding="utf-8")
    archivo_temporal.replace(OUTPUT_PATH)


def main() -> None:
    try:
        deportes = leer_deportes()
        escribir_javascript(deportes)

        activos = sum(1 for deporte in deportes if deporte["activo"])
        permiten_eventos = sum(
            1 for deporte in deportes if deporte["permiteEventos"]
        )

        print("✅ deportes.js actualizado correctamente.")
        print(f"   Deportes totales: {len(deportes)}")
        print(f"   Activos: {activos}")
        print(f"   Permiten eventos: {permiten_eventos}")
        print(f"   Archivo creado: {OUTPUT_PATH}")

        backup_path = OUTPUT_PATH.with_suffix(".js.bak")

        if backup_path.exists():
            print(f"   Copia anterior: {backup_path}")

    except Exception as error:
        print("\n❌ No se pudo actualizar deportes.js:")
        print(f"   {error}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
